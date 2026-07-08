"""End-to-end tests for write_output.build_output: merges synthetic research
results onto a synthetic RPA-Products-shaped workbook and checks the
resulting 'nieuw' columns, per references/test-plan.md section 1 and the
skill's own decision rules.

Covers rules 1/2/3/5 in a full workbook round-trip, plus rule 6 (rerun with
--previous-output correctly preserves a prior 'handmatig aanpassen'
decision instead of letting a fresh proposal silently overwrite it).
"""
from __future__ import annotations

import json

import openpyxl
import schema
from write_output import build_output


def _make_results(rows: list[dict]) -> dict:
    return {"rows": rows}


def _read_output_row(output_path, code_huidig: str) -> dict:
    wb = openpyxl.load_workbook(output_path, data_only=True)
    ws = wb["Verrijking"]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["code (huidig)"]] == code_huidig:
            # openpyxl round-trips a written "" as None on reload; normalize
            # back to "" so assertions can compare against decision_logic's
            # actual return values instead of this storage quirk.
            return {h: ("" if row[i] is None else row[i]) for h, i in idx.items()}
    raise AssertionError(f"row with code (huidig)={code_huidig!r} not found in output")


def test_full_workbook_covers_rules_1_2_3_5(tmp_path, rpa_workbook_factory, sample_row):
    rows = [
        # Rule 1: huidig empty + Hoog confidence -> auto-fill
        sample_row({"naam": "Metacam 1.5mg/ml", "code": "METACAM15", "registratie": ""}),
        # Rule 2: huidig == voorstel -> unchanged
        sample_row({"naam": "Rimadyl 50mg", "code": "RIMADYL50", "registratie": "REG-NL-1111"}),
        # Rule 3: huidig != voorstel -> stays huidig, flagged for review
        sample_row({"naam": "Clavubactin 500/125mg", "code": "CLAVU500", "registratie": "REG-NL-OLD"}),
        # Rule 5: no result found -> nieuw = huidig (already filled) / stays empty
        sample_row({"naam": "Fictief Testproduct XYZ", "code": "FICTIEF1", "registratie": "REG-NL-KEEP"}),
        sample_row({"naam": "Fictief Testproduct Leeg", "code": "FICTIEF2", "registratie": ""}),
    ]
    input_path = rpa_workbook_factory(rows=rows)

    results = _make_results([
        {"row_index": 2, "fields": {"registratie": {"voorstel": "REG-NL-9999", "confidence": "Hoog"}},
         "resultaatstatus": "exact gevonden"},
        {"row_index": 3, "fields": {"registratie": {"voorstel": "REG-NL-1111", "confidence": "Hoog"}},
         "resultaatstatus": "exact gevonden"},
        {"row_index": 4, "fields": {"registratie": {"voorstel": "REG-NL-NEW", "confidence": "Hoog"}},
         "resultaatstatus": "handmatige controle nodig"},
        {"row_index": 5, "fields": {}, "resultaatstatus": "niet gevonden"},
        {"row_index": 6, "fields": {}, "resultaatstatus": "niet gevonden"},
    ])
    results_path = tmp_path / "results.json"
    results_path.write_text(json.dumps(results), encoding="utf-8")

    output_path = tmp_path / "output.xlsx"
    summary = build_output(str(input_path), str(results_path), str(output_path))

    # Rule 1
    row = _read_output_row(output_path, "METACAM15")
    assert row["registratie (nieuw)"] == "REG-NL-9999"

    # Rule 2
    row = _read_output_row(output_path, "RIMADYL50")
    assert row["registratie (nieuw)"] == "REG-NL-1111"

    # Rule 3: huidig wins, never auto-overwritten, and the row counts as
    # needing review (both via decision_logic's flag and resultaatstatus).
    row = _read_output_row(output_path, "CLAVU500")
    assert row["registratie (nieuw)"] == "REG-NL-OLD"
    assert row["resultaatstatus"] == "handmatige controle nodig"

    # Rule 5: huidig already filled, no result found -> nieuw = huidig
    row = _read_output_row(output_path, "FICTIEF1")
    assert row["registratie (nieuw)"] == "REG-NL-KEEP"

    # Rule 5: huidig empty, no result found -> nieuw stays empty
    row = _read_output_row(output_path, "FICTIEF2")
    assert row["registratie (nieuw)"] == ""

    assert summary["rows_written"] == 5
    assert summary["review_rows"] >= 1

    # Reopening must not raise (write_output.py's own smoke test already
    # does this, but re-asserting here documents the guarantee).
    openpyxl.load_workbook(output_path)


def test_gemiddeld_confidence_on_empty_huidig_never_auto_adopted(tmp_path, rpa_workbook_factory, sample_row):
    rows = [sample_row({"naam": "Onzeker Product", "code": "ONZ1", "kanalisatie": ""})]
    input_path = rpa_workbook_factory(rows=rows)
    results = _make_results([
        {"row_index": 2, "fields": {"kanalisatie": {"voorstel": "uda", "confidence": "Gemiddeld"}},
         "resultaatstatus": "gedeeltelijk verrijkt"},
    ])
    results_path = tmp_path / "results.json"
    results_path.write_text(json.dumps(results), encoding="utf-8")
    output_path = tmp_path / "output.xlsx"

    build_output(str(input_path), str(results_path), str(output_path))

    row = _read_output_row(output_path, "ONZ1")
    assert row["kanalisatie (nieuw)"] == ""


# --- Rule 6: rerun with --previous-output preserves 'handmatig aanpassen' -

def test_rerun_preserves_handmatig_aanpassen_and_does_not_let_new_proposal_overwrite(
    tmp_path, rpa_workbook_factory, sample_row
):
    input_path = rpa_workbook_factory(rows=[
        sample_row({"naam": "Cerenia 16mg", "code": "CERENIA16", "registratie": ""}),
    ])

    # --- run 1: an initial (say, low-confidence / conflicting) proposal ---
    results_v1 = _make_results([
        {"row_index": 2, "fields": {"registratie": {"voorstel": "REG-NL-CANDIDATE-A", "confidence": "Gemiddeld"}},
         "resultaatstatus": "handmatige controle nodig"},
    ])
    results_v1_path = tmp_path / "results_v1.json"
    results_v1_path.write_text(json.dumps(results_v1), encoding="utf-8")
    output_v1_path = tmp_path / "output_v1.xlsx"
    build_output(str(input_path), str(results_v1_path), str(output_v1_path))

    # A human then reviews output_v1 and decides to manually set the value
    # themselves ("handmatig aanpassen"), entering their own chosen nieuw.
    wb = openpyxl.load_workbook(output_v1_path)
    ws = wb["Verrijking"]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    for row in ws.iter_rows(min_row=2):
        if row[idx["code (huidig)"]].value == "CERENIA16":
            row[idx["beslissing (registratie)"]].value = "handmatig aanpassen"
            row[idx["registratie (nieuw)"]].value = "REG-NL-USER-CHOSEN"
    wb.save(output_v1_path)

    # --- run 2: a fresh, high-confidence, DIFFERENT proposal shows up -----
    results_v2 = _make_results([
        {"row_index": 2, "fields": {"registratie": {"voorstel": "REG-NL-BRAND-NEW-HOOG", "confidence": "Hoog"}},
         "resultaatstatus": "exact gevonden"},
    ])
    results_v2_path = tmp_path / "results_v2.json"
    results_v2_path.write_text(json.dumps(results_v2), encoding="utf-8")
    output_v2_path = tmp_path / "output_v2.xlsx"

    build_output(
        str(input_path), str(results_v2_path), str(output_v2_path),
        previous_output_path=str(output_v1_path),
    )

    row = _read_output_row(output_v2_path, "CERENIA16")
    # the user's manual decision must survive, NOT the new Hoog proposal
    assert row["registratie (nieuw)"] == "REG-NL-USER-CHOSEN"
    assert row["beslissing (registratie)"] == "handmatig aanpassen"


def test_rerun_akkoord_decision_adopts_latest_voorstel(tmp_path, rpa_workbook_factory, sample_row):
    input_path = rpa_workbook_factory(rows=[
        sample_row({"naam": "Previcox 57mg", "code": "PREVICOX57", "registratie": ""}),
    ])

    results_v1 = _make_results([
        {"row_index": 2, "fields": {"registratie": {"voorstel": "REG-NL-V1", "confidence": "Gemiddeld"}},
         "resultaatstatus": "handmatige controle nodig"},
    ])
    results_v1_path = tmp_path / "results_v1.json"
    results_v1_path.write_text(json.dumps(results_v1), encoding="utf-8")
    output_v1_path = tmp_path / "output_v1.xlsx"
    build_output(str(input_path), str(results_v1_path), str(output_v1_path))

    # Reviewer marks it "akkoord" in the v1 output.
    wb = openpyxl.load_workbook(output_v1_path)
    ws = wb["Verrijking"]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    for row in ws.iter_rows(min_row=2):
        if row[idx["code (huidig)"]].value == "PREVICOX57":
            row[idx["beslissing (registratie)"]].value = "akkoord"
    wb.save(output_v1_path)

    # Run 2 finds an updated/refined proposal.
    results_v2 = _make_results([
        {"row_index": 2, "fields": {"registratie": {"voorstel": "REG-NL-V2-REFINED", "confidence": "Hoog"}},
         "resultaatstatus": "exact gevonden"},
    ])
    results_v2_path = tmp_path / "results_v2.json"
    results_v2_path.write_text(json.dumps(results_v2), encoding="utf-8")
    output_v2_path = tmp_path / "output_v2.xlsx"

    build_output(
        str(input_path), str(results_v2_path), str(output_v2_path),
        previous_output_path=str(output_v1_path),
    )

    row = _read_output_row(output_v2_path, "PREVICOX57")
    # "akkoord" recomputes with the CURRENT run's voorstel (not frozen at v1)
    assert row["registratie (nieuw)"] == "REG-NL-V2-REFINED"
    assert row["beslissing (registratie)"] == "akkoord"


def test_reopening_output_workbook_does_not_raise(tmp_path, rpa_workbook_factory, sample_row):
    input_path = rpa_workbook_factory(rows=[sample_row({"naam": "X", "code": "X1"})])
    results_path = tmp_path / "results.json"
    results_path.write_text(json.dumps(_make_results([
        {"row_index": 2, "fields": {}, "resultaatstatus": "niet gevonden"},
    ])), encoding="utf-8")
    output_path = tmp_path / "output.xlsx"

    build_output(str(input_path), str(results_path), str(output_path))

    wb = openpyxl.load_workbook(output_path)
    assert "Verrijking" in wb.sheetnames
    assert "Verwerkingsrapport" in wb.sheetnames
