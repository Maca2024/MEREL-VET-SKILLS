"""Step 3 of the skill: merge Claude's research results back onto the
original rows and write the final enriched workbook.

Design note (important): the input workbook is opened READ-ONLY for values
only and the output is built as a brand-new openpyxl Workbook. We do NOT
load-then-save the original worksheet object. The RPA Products template
carries an Excel data-validation extension that openpyxl explicitly does not
round-trip ("Data Validation extension is not supported and will be
removed" -- confirmed against dev-fixtures/Templete RPA Products (1).xlsx).
Copying that worksheet object and re-saving it risks a "needs repair" prompt
on the client's machine. Building fresh avoids that risk entirely, and this
output workbook is a review artifact, not something re-imported into
Animana, so it doesn't need the original's dropdown validations anyway.

Usage:
    python write_output.py <input.xlsx> <enrichment_results.json> <output.xlsx> [--previous-output prior.xlsx]

Never modifies the input file in place.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from schema import (
    RPA_COLUMNS, REQUIRED_COLUMNS, STANDARD_FIELDS, CODE_OUTPUT_COLUMNS,
    ROW_CONTROL_COLUMNS, full_output_columns, field_columns, waarschuwing_columns,
    TEXT_FORMAT_COLUMNS, RESULTAATSTATUS_VALUES,
)
from decision_logic import compute_nieuw, compose_code_plus_zoektermen, compute_code_nieuw
from validate_and_export import find_product_sheet, _norm_header


def _read_original_rows(input_path: str):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = find_product_sheet(wb)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_index = {_norm_header(v): i for i, v in enumerate(header_row) if v is not None}

    rows = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or v == "" for v in row):
            continue
        record = {}
        for col in RPA_COLUMNS:
            idx = header_index.get(col.lower())
            value = row[idx] if idx is not None and idx < len(row) else None
            record[col] = "" if value is None else value
        rows[row_idx] = record
    return rows


def _load_previous_decisions(previous_output_path: str | None):
    """Re-run support: pull forward beslissing + (for handmatig aanpassen)
    the user's own nieuw value, keyed by 'code' (falls back to row order if
    code is blank/duplicated -- documented limitation, see SKILL.md)."""
    if not previous_output_path:
        return {}
    wb = openpyxl.load_workbook(previous_output_path, data_only=True)
    ws = wb["Verrijking"] if "Verrijking" in wb.sheetnames else wb.worksheets[0]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_index = {h: i for i, h in enumerate(header_row) if h is not None}

    prior = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code_huidig_idx = header_index.get("code (huidig)")
        if code_huidig_idx is None:
            continue
        key = row[code_huidig_idx]
        if key in (None, ""):
            continue
        entry = {"beslissing": {}, "nieuw": {}}
        for field in STANDARD_FIELDS + ["waarschuwing", "code"]:
            b_col, n_col = f"beslissing ({field})", f"{field} (nieuw)"
            if field == "code":
                b_col, n_col = "beslissing (code)", "code (nieuw)"
            if b_col in header_index:
                entry["beslissing"][field] = row[header_index[b_col]]
            if n_col in header_index:
                entry["nieuw"][field] = row[header_index[n_col]]
        prior[key] = entry
    return prior


def build_output(input_path: str, results_path: str, output_path: str, previous_output_path: str | None = None):
    original_rows = _read_original_rows(input_path)
    results = json.loads(Path(results_path).read_text(encoding="utf-8"))
    results_by_row = {r["row_index"]: r for r in results.get("rows", results if isinstance(results, list) else [])}
    prior_decisions = _load_previous_decisions(previous_output_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verrijking"

    columns = full_output_columns()
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    stats = {v: 0 for v in RESULTAATSTATUS_VALUES}
    review_rows = 0

    for row_idx in sorted(original_rows.keys()):
        orig = original_rows[row_idx]
        res = results_by_row.get(row_idx, {})
        fields = res.get("fields", {})
        code_huidig = orig.get("code", "")
        prior = prior_decisions.get(code_huidig, {}) if code_huidig else {}

        row_out = {col: orig.get(col, "") for col in RPA_COLUMNS}
        row_needs_review = False

        # --- code block ---
        zoektermen = fields.get("zoektermen", {}).get("voorstel", []) if isinstance(fields.get("zoektermen"), dict) else fields.get("zoektermen", [])
        code_plus_terms = compose_code_plus_zoektermen(code_huidig, zoektermen)
        beslissing_code = prior.get("beslissing", {}).get("code") or ""
        code_nieuw, _ = compute_code_nieuw(code_huidig, code_plus_terms, beslissing_code)
        if code_nieuw is None:  # handmatig aanpassen sentinel
            code_nieuw = prior.get("nieuw", {}).get("code", "")
        row_out["code (huidig)"] = code_huidig
        row_out["zoektermen (voorstel)"] = ", ".join(zoektermen) if zoektermen else ""
        row_out["code + zoektermen (voorstel)"] = code_plus_terms
        row_out["beslissing (code)"] = beslissing_code
        row_out["code (nieuw)"] = code_nieuw

        # --- waarschuwing (special: has bronpassage) ---
        w = fields.get("waarschuwing", {})
        huidig_w = orig.get("waarschuwing", "")
        voorstel_w = w.get("voorstel", "")
        conf_w = w.get("confidence", "Niet gevonden")
        beslissing_w = prior.get("beslissing", {}).get("waarschuwing") or ""
        nieuw_w, flag_w = compute_nieuw(huidig_w, voorstel_w, conf_w, beslissing_w)
        if nieuw_w is None:
            nieuw_w = prior.get("nieuw", {}).get("waarschuwing", "")
        row_out["waarschuwing (huidig)"] = huidig_w
        row_out["waarschuwing (voorstel)"] = voorstel_w
        row_out["waarschuwing (confidence voorstel)"] = conf_w
        row_out["waarschuwing bronpassage (voorstel)"] = w.get("bronpassage", "")
        row_out["waarschuwing (nieuw)"] = nieuw_w
        row_out["beslissing (waarschuwing)"] = beslissing_w
        row_needs_review = row_needs_review or flag_w

        # --- standard fields ---
        for field in STANDARD_FIELDS:
            f = fields.get(field, {})
            huidig_v = orig.get(field, "")
            voorstel_v = f.get("voorstel", "")
            conf_v = f.get("confidence", "Niet gevonden")
            beslissing_v = prior.get("beslissing", {}).get(field) or ""
            nieuw_v, flag_v = compute_nieuw(huidig_v, voorstel_v, conf_v, beslissing_v)
            if nieuw_v is None:
                nieuw_v = prior.get("nieuw", {}).get(field, "")
            row_out[f"{field} (huidig)"] = huidig_v
            row_out[f"{field} (voorstel)"] = voorstel_v
            row_out[f"{field} (confidence voorstel)"] = conf_v
            row_out[f"{field} (nieuw)"] = nieuw_v
            row_out[f"beslissing ({field})"] = beslissing_v
            row_needs_review = row_needs_review or flag_v

        # --- row-level control columns ---
        resultaatstatus = res.get("resultaatstatus", "niet gevonden")
        if resultaatstatus not in stats:
            resultaatstatus = "handmatige controle nodig"
        stats[resultaatstatus] += 1
        if row_needs_review or resultaatstatus in ("meerdere mogelijke matches", "handmatige controle nodig"):
            review_rows += 1

        row_out["bron URL"] = res.get("bron_url", "")
        row_out["resultaatstatus"] = resultaatstatus
        row_out["toelichting controle"] = res.get("toelichting_controle", "")

        ws.append([row_out[col] for col in columns])

    # barcode-like columns as text, preserve leading zeros
    text_col_letters = [get_column_letter(i + 1) for i, c in enumerate(columns) if c in TEXT_FORMAT_COLUMNS]
    for letter in text_col_letters:
        for cell in ws[letter][1:]:
            cell.number_format = "@"

    # sensible fixed column widths + autofilter
    ws.auto_filter.ref = ws.dimensions
    for i, col in enumerate(columns, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = min(max(len(col) + 2, 12), 40)

    _write_report_sheet(wb, stats, len(original_rows), review_rows)

    wb.save(output_path)

    # smoke test: reopen the file we just wrote to catch corruption early
    openpyxl.load_workbook(output_path)
    return {"rows_written": len(original_rows), "stats": stats, "review_rows": review_rows}


def _write_report_sheet(wb, stats: dict, total_rows: int, review_rows: int):
    rs = wb.create_sheet("Verwerkingsrapport")
    rs.append(["Verwerkingsrapport", ""])
    rs["A1"].font = Font(bold=True, size=14)
    rs.append(["Totaal producten", total_rows])
    usable = total_rows - stats.get("niet gevonden", 0)
    coverage = (usable / total_rows * 100) if total_rows else 0
    rs.append(["Bruikbaar resultaat (coverage)", f"{coverage:.1f}%"])
    rs.append(["Rijen die handmatige review nodig hebben", review_rows])
    rs.append([])
    rs.append(["resultaatstatus", "aantal"])
    rs["A6"].font = Font(bold=True)
    rs["B6"].font = Font(bold=True)
    for status, count in stats.items():
        rs.append([status, count])
    for col, width in (("A", 42), ("B", 14)):
        rs.column_dimensions[col].width = width


def main():
    p = argparse.ArgumentParser()
    p.add_argument("input_xlsx")
    p.add_argument("results_json")
    p.add_argument("output_xlsx")
    p.add_argument("--previous-output")
    args = p.parse_args()

    summary = build_output(args.input_xlsx, args.results_json, args.output_xlsx, args.previous_output)
    print(f"OK: wrote {args.output_xlsx}")
    print(f"  rows written: {summary['rows_written']}")
    print(f"  rows flagged for review: {summary['review_rows']}")
    print("  resultaatstatus breakdown:")
    for status, count in summary["stats"].items():
        if count:
            print(f"    {status}: {count}")


if __name__ == "__main__":
    main()
