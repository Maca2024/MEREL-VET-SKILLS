"""Step 1 of the skill: validate an RPA-Products-shaped input workbook and
export its rows to JSON so Claude can run the research loop without touching
openpyxl / worksheet formatting directly.

This script is deliberately independent of Skill 1: it makes NO assumption
that the input came fresh from that skill. It only requires that the header
row contains the columns this skill needs (schema.REQUIRED_COLUMNS). If any
are missing, it reports exactly which ones and exits non-zero -- it never
guesses a substitute column.

Usage:
    python validate_and_export.py <input.xlsx> <output.json>

Exit codes:
    0  ok, JSON written
    1  input file / sheet problem
    2  required columns missing (message lists exactly which)
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

from schema import REQUIRED_COLUMNS


def _norm_header(value) -> str:
    return "" if value is None else str(value).strip().lower()


def find_product_sheet(wb):
    """Find the worksheet whose header row contains the required columns.
    Robust to a renamed sheet (not strictly "RPA Products edit") because a
    hand-edited or older-run file may not use that exact sheet name."""
    required_norm = {c.lower() for c in REQUIRED_COLUMNS}
    best_sheet, best_hits = None, -1
    for ws in wb.worksheets:
        header = [_norm_header(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
        hits = len(required_norm & set(header))
        if hits > best_hits:
            best_sheet, best_hits = ws, hits
    return best_sheet


def validate_and_export(input_path: str, output_path: str) -> int:
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = find_product_sheet(wb)
    if ws is None:
        print("ERROR: workbook has no worksheets.")
        return 1

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = [_norm_header(c) for c in header_row]
    header_index = {name: idx for idx, name in enumerate(header)}

    missing = [c for c in REQUIRED_COLUMNS if c.lower() not in header_index]
    if missing:
        print("ERROR: input workbook is missing required column(s):")
        for m in missing:
            print(f"  - {m}")
        print("Stopping. Fix the input file's header row and re-run "
              "(no substitute column will be guessed).")
        return 2

    # Preserve original casing of the columns as they appear in the file,
    # since write_output.py must be able to read the ORIGINAL workbook again
    # using the same header names.
    original_names = {_norm_header(v): v for v in header_row if v is not None}

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or v == "" for v in row):
            continue
        record = {"row_index": row_idx}
        for col in REQUIRED_COLUMNS:
            idx = header_index[col.lower()]
            value = row[idx] if idx < len(row) else None
            record[col] = "" if value is None else value
        rows.append(record)

    payload = {
        "source_file": str(Path(input_path).resolve()),
        "sheet_name": ws.title,
        "original_header_casing": original_names,
        "row_count": len(rows),
        "rows": rows,
    }
    Path(output_path).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK: {len(rows)} product row(s) exported from sheet '{ws.title}' -> {output_path}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python validate_and_export.py <input.xlsx> <output.json>")
        sys.exit(1)
    sys.exit(validate_and_export(sys.argv[1], sys.argv[2]))
