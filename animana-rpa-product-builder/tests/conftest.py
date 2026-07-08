"""Shared pytest fixtures for the Animana RPA Product Builder skill.

Adds the skill's scripts/ directory to sys.path so tests can `import
rpa_builder...` the same way build_rpa_products.py does, and provides a
`run_cli()` helper that invokes the real CLI (scripts/build_rpa_products.py)
as a subprocess -- the same way an operator would run it -- rather than
calling internal functions directly. This is deliberate: several of the
required assertions (exact output file count, collision-avoidance naming,
byte-identical template passthrough) are only meaningful if exercised
end-to-end through main().
"""
from __future__ import annotations

import subprocess
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pytest

SKILL_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS_DIR = SKILL_ROOT / "scripts"
BUILD_SCRIPT = SCRIPTS_DIR / "build_rpa_products.py"
TEMPLATE_PATH = SKILL_ROOT / "assets" / "RPA_Products_template.xlsx"
DEV_FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures" / "dev_fixtures"

if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from rpa_builder import xlsx_xml as xx  # noqa: E402


@dataclass
class CliResult:
    returncode: int
    stdout: str
    stderr: str
    output_dir: Path

    def output_files(self) -> list[Path]:
        return sorted(p for p in self.output_dir.iterdir() if p.is_file())

    def xlsx_files(self) -> list[Path]:
        return [p for p in self.output_files() if p.suffix == ".xlsx"]


def run_cli(output_dir: Path, **kwargs) -> CliResult:
    """Runs the real CLI as a subprocess. kwargs map argparse dest names
    (underscored) to values; Path values are passed as strings. Booleans/None
    are skipped. `account_name` defaults to a fixed test account name and
    `output_dir` is always kwargs['output_dir'] unless overridden."""
    args = [sys.executable, str(BUILD_SCRIPT)]
    kwargs.setdefault("account_name", "Test Kliniek Dev")
    kwargs.setdefault("output_dir", output_dir)
    for key, value in kwargs.items():
        if value is None:
            continue
        flag = "--" + key.replace("_", "-")
        args.append(flag)
        args.append(str(value))
    proc = subprocess.run(args, capture_output=True, text=True, encoding="utf-8", errors="replace")
    return CliResult(proc.returncode, proc.stdout, proc.stderr, Path(output_dir))


def run_full_dev_fixtures(output_dir: Path, **overrides) -> CliResult:
    """Runs the CLI against the full anonymized dev-fixtures set (all
    optional inputs provided), unless a test overrides a specific kwarg
    (e.g. to point --product-export at a synthetic file instead)."""
    kwargs = dict(
        product_export=DEV_FIXTURES_DIR / "export_product_dev.csv",
        productgroups=DEV_FIXTURES_DIR / "productgroups.xlsx",
        contacts=DEV_FIXTURES_DIR / "export_client_dev.csv",
        letters=DEV_FIXTURES_DIR / "export.xlsx",
        toeslagen=DEV_FIXTURES_DIR / "Voorbeeld copy-paste toeslagen.txt",
        margeregels=DEV_FIXTURES_DIR / "Voorbeeld copy-paste margeregels.txt",
    )
    kwargs.update(overrides)
    return run_cli(output_dir, **kwargs)


def load_sheet1_rows(xlsx_path: Path) -> list[list]:
    """Header + data rows of 'RPA Products edit' as plain values (openpyxl,
    read-only). Column A is index 0."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["RPA Products edit"]
        return [[c.value for c in row] for row in ws.iter_rows()]
    finally:
        wb.close()


def read_zip_part(xlsx_path: Path, part: str) -> str:
    with zipfile.ZipFile(xlsx_path) as z:
        return z.read(part).decode("utf-8")


@pytest.fixture
def cli(tmp_path):
    """Callable: cli(**kwargs) -> CliResult, output_dir defaults to a fresh
    subdirectory of tmp_path per call unless output_dir is passed explicitly."""
    counter = {"n": 0}

    def _run(**kwargs):
        if "output_dir" not in kwargs:
            counter["n"] += 1
            out = tmp_path / f"out{counter['n']}"
            out.mkdir(exist_ok=True)
            kwargs["output_dir"] = out
        return run_cli(**kwargs)

    return _run


@pytest.fixture
def minimal_product_header() -> list[str]:
    """The 7 columns readers.PRODUCT_EXPORT_SIGNATURE requires, plus a couple
    of commonly-mapped ones used by synthetic-fixture tests. Any DIRECT_FIELD_MAP
    source column not present here is simply read as '' by readers.py -- no
    error -- so this keeps synthetic CSVs small and focused."""
    return [
        "id", "productgroep", "naam", "verkoopprijs", "kostprijs", "actief",
        "barcode", "kanalisatie", "shortcode", "btw",
    ]


def write_minimal_productgroups(path: Path, names: list[str] | None = None) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["naam", "producten", "margeregel", "prijsgroep"])
    for n in names or ["Groep A"]:
        ws.append([n, "", "", ""])
    wb.save(path)
    return path
