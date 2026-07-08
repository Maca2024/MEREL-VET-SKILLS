"""Regression test for a real bug found via a live smoke test on the
Kathedraal (8 jul 2026): --input-dir auto-classification swept a plain-prose
LEESMIJ/README .txt file into the 'toeslagen' dropdown (column C) because
classify_copy_paste_file() had no sanity check for "does this even look like
a tab-separated copy-paste export" -- a prose file with no tabs defaulted to
'toeslagen' by the fallback branch, and (being alphabetically first) beat the
real 'Voorbeeld copy-paste toeslagen.txt' file to the slot. Confirmed live:
the shipped Skill 1's real dev account export produced a 'dropdowns' column C
containing README prose instead of real toeslagen names.
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))

from rpa_builder import readers
from conftest import DEV_FIXTURES_DIR, write_minimal_productgroups, minimal_product_header  # noqa: F401


def _write(path: Path, text: str) -> Path:
    path.write_text(text, encoding="utf-8")
    return path


def test_classify_returns_none_for_prose_readme(tmp_path):
    readme = _write(tmp_path / "LEESMIJ_v2.txt", (
        "AetherLink overdrachtspakket - Skill 1, versie 2\n\n"
        "Belangrijk uitgangspunt:\n"
        "Alle exports en ingevulde bestanden in dit pakket zijn uitsluitend "
        "voorbeelden van de structuur.\n"
    ))
    assert readers.classify_copy_paste_file(readme) is None


def test_classify_still_recognizes_real_toeslagen_and_margeregels(tmp_path):
    toeslagen = _write(tmp_path / "toeslagen.txt", "Injectie\t€ 12,00\t\nRecept - chronisch\t€ 3,00\t\n")
    margeregels = _write(tmp_path / "margeregels.txt", "Basisprijs factor 1,5\t€ 0,00\t€ 0,00\t\n")
    assert readers.classify_copy_paste_file(toeslagen) == "toeslagen"
    assert readers.classify_copy_paste_file(margeregels) == "margeregels"


def test_input_dir_end_to_end_ignores_readme_and_finds_real_toeslagen(cli, tmp_path):
    """End-to-end: build an --input-dir containing a LEESMIJ.txt (alphabetically
    first) plus the real toeslagen/margeregels files, and confirm the CLI picks
    the real toeslagen file, not the readme."""
    input_dir = tmp_path / "client-exports"
    input_dir.mkdir()
    _write(input_dir / "AAA_LEESMIJ.txt", (
        "Dit is een leesmij-bestand met uitleg over het pakket.\n"
        "Geen kolommen, geen tabs, gewoon lopende tekst.\n"
    ))
    (input_dir / "export_product.csv").write_bytes(
        (DEV_FIXTURES_DIR / "export_product_dev.csv").read_bytes())
    write_minimal_productgroups(input_dir / "productgroups.xlsx")
    (input_dir / "toeslagen.txt").write_text(
        "Injectie\t€ 12,00\t\nRecept - chronisch\t€ 3,00\t\n", encoding="utf-8")

    result = cli(input_dir=input_dir)
    assert result.returncode == 0, result.stdout + result.stderr

    import openpyxl
    xlsx = next(p for p in result.xlsx_files() if "compleet" in p.name)
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["dropdowns"]
    col_c_values = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=3).value]
    assert "Injectie" in col_c_values, f"expected real toeslagen values, got: {col_c_values[:5]}"
    assert not any("leesmij" in str(v).lower() or "AetherLink overdrachtspakket" in str(v) for v in col_c_values), (
        f"README prose leaked into the toeslagen dropdown: {col_c_values[:5]}")
